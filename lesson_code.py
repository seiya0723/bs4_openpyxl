## ドル円相場を取得して、エクセルファイルに日時もセットで追記する

import requests,bs4
import openpyxl as px

import datetime

URL     = "https://www.nikkei.com/markets/kawase/"
data    = requests.get(URL)
soup    = bs4.BeautifulSoup(data.content , "html.parser")


elems   = soup.select(".mkc-stock_prices")

for elem in elems:
    print("現在のレート")
    print(elem.text)

    raw_text    = elem.text

    #-で区切ってリスト型に直す(※ただ、これだとスペースが混ざってしまう)
    low_high    = raw_text.split("-")

    #スペースを消した上で-で区切る
    low_high    = raw_text.replace(" ","").split("-")


    #エクセルファイルを読み込み
    wb          = px.load_workbook("test.xlsx")

    #アクティブシートを選択(新規作成時に最初からあるシート)
    ws          = wb.active

    #シートの最終行の次の行を取得
    max_row     = ws.max_row + 1

    #最終行のA列からC列まで順に値を入れる
    ws.cell(row=max_row, column=1).value    = datetime.datetime.now()
    ws.cell(row=max_row, column=2).value    = float(low_high[0])
    ws.cell(row=max_row, column=3).value    = float(low_high[1])

    wb.save("test.xlsx")

